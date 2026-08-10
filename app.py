#!/usr/bin/env python3
"""Casa Nova — local web app: cash entry, bank sync, analytics."""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response
import openpyxl
import urllib.request
import json
import subprocess
import socket
import sys
import os
import re
import qrcode
from datetime import datetime, date
from collections import defaultdict
import time

app = Flask(__name__)
app.secret_key = 'casanova-local-2025'  # local-only, no auth — rotate if app ever goes public

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
NOVA_FILE    = os.path.join(BASE_DIR, 'Custos Casa Nova - Nova2.xlsx')
DASHBOARD_PY = os.path.join(BASE_DIR, 'dashboard.py')
STATUS_FILE  = os.path.join(BASE_DIR, 'category_status.json')
TOKEN_FILE   = os.path.expanduser('~/Documents/Personal/Finance/ynab token.rtf')
ANTHROPIC_KEY_FILE = os.path.expanduser('~/Documents/Personal/Finance/anthropic api key.txt')

def load_category_status():
    if not os.path.exists(STATUS_FILE): return {}
    try:
        with open(STATUS_FILE) as f: return json.load(f)
    except Exception: return {}

def save_category_status(s):
    with open(STATUS_FILE, 'w') as f: json.dump(s, f, indent=2, ensure_ascii=False)

def load_ynab_token():
    """Read the YNAB personal access token from the RTF file at TOKEN_FILE.
    Tokens are 43 chars of URL-safe base64. We grep the RTF for the first match."""
    try:
        with open(TOKEN_FILE) as f:
            content = f.read()
    except FileNotFoundError:
        raise RuntimeError(
            f'YNAB token file not found at {TOKEN_FILE}. '
            'Create it with your token (see README).'
        )
    m = re.search(r'(?<![A-Za-z0-9_-])[A-Za-z0-9_-]{43}(?![A-Za-z0-9_-])', content)
    if not m:
        raise RuntimeError(f'Could not find a valid YNAB token in {TOKEN_FILE}')
    return m.group(0)

TOKEN        = load_ynab_token()
BUDGET_ID    = '62da99f2-0125-45cd-8906-6a5ebe3416ad'
CAT_ID       = '6bc4c34f-60e3-462b-9f75-a9baf10d35b0'
CASH_OBRA_ID = '696f7b16-febb-4f9b-93e0-301a591c297e'
MG_ID        = 'd2e9c80d-822f-4d22-9666-ce79a3fc41ae'

BASE_TAGS = [
    'Terreno', 'Arquitectura', 'Construção grosso', 'Pladur',
    'Pichelaria & Climatização', 'Aluminio', 'Carpintaria',
    'Construção acabamentos', 'Capoto', 'Electricidade', 'Pintura',
    'Piscina', 'Estores', 'Furo', 'Design & Decoração Interiores',
    'Serralheiro', 'Loiças & Cerâmicos', 'Electrodomésticos',
    'Agua & Luz', 'Outros', 'Taxas & Multas', 'Empréstimo'
]

BUDGET = {
    'Terreno': 70_000, 'Arquitectura': 4_500, 'Construção grosso': 80_000,
    'Pladur': 22_000, 'Pichelaria & Climatização': 25_000, 'Aluminio': 25_000,
    'Carpintaria': 40_000, 'Construção acabamentos': 30_000, 'Capoto': 15_000,
    'Electricidade': 15_000, 'Pintura': 15_000, 'Piscina': 18_000,
    'Estores': 7_000, 'Furo': 5_000, 'Design & Decoração Interiores': 10_400,
    'Serralheiro': 10_000, 'Loiças & Cerâmicos': 8_000, 'Electrodomésticos': 10_000,
    'Agua & Luz': 1_000, 'Outros': 10_000,
}
IVA_RATE   = 0.23
IVA_EXEMPT = {'Terreno'}

def budget_civa(tag):
    b = BUDGET.get(tag, 0)
    return b if tag in IVA_EXEMPT else round(b * (1 + IVA_RATE), 2)

TOTAL_CIVA = sum(budget_civa(t) for t in BUDGET)
TOTAL_SIVA = sum(BUDGET.values())

# ── YNAB helpers ──────────────────────────────────────────────────────────────
def ynab_get(path):
    req = urllib.request.Request(
        f'https://api.ynab.com/v1{path}',
        headers={'Authorization': f'Bearer {TOKEN}'}
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read())

def ynab_post(path, body):
    data = json.dumps(body).encode()
    req = urllib.request.Request(
        f'https://api.ynab.com/v1{path}', data=data,
        headers={'Authorization': f'Bearer {TOKEN}', 'Content-Type': 'application/json'}
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read())

# ── Liquidity source of truth: YNAB live, with Parâmetros as fallback ────────
# "Disponibilidade Casa Nova" is the YNAB *category* available balance for the
# build — money allocated and ready to spend on the house, independent of what
# else sits in the same bank account. This is the canonical headline number.
# Bank + Cash account balances are also fetched, as a secondary "where does
# the money sit" breakdown. Every expense logged via the app posts to YNAB, so
# both numbers update on next read. If YNAB is unreachable we fall back to
# B2/B3 from the Parâmetros sheet, and the UI shows ⚠ fallback so the staleness
# is never silent.
_balance_cache = {'data': None, 'ts': 0.0}
BALANCE_TTL = 60  # seconds — short enough to feel live, long enough to avoid hammering YNAB

def ynab_snapshot():
    """Return (available_eur, bank_eur, cash_eur, source).

    available: YNAB 'available' for the Casa Nova category (CAT_ID) — the
      true disponibilidade for the build.
    bank, cash: balances of MG Ordenado and Cache obra accounts — shown as
      breakdown context, not the headline.
    source: 'live' when fetched from YNAB, 'fallback' when YNAB was
      unreachable. In fallback we approximate available as bank + cash from
      the Parâmetros sheet, which is a rougher proxy.
    """
    now = time.time()
    cached = _balance_cache['data']
    if cached and (now - _balance_cache['ts'] < BALANCE_TTL):
        return cached
    try:
        cat = ynab_get(f'/budgets/{BUDGET_ID}/categories/{CAT_ID}')
        available = cat['data']['category']['balance'] / 1000.0
        accs = ynab_get(f'/budgets/{BUDGET_ID}/accounts')
        accounts = {a['id']: a for a in accs['data']['accounts'] if not a.get('deleted')}
        bank = accounts[MG_ID]['balance'] / 1000.0
        cash = accounts[CASH_OBRA_ID]['balance'] / 1000.0
        result = (available, bank, cash, 'live')
        _balance_cache.update(data=result, ts=now)
        return result
    except Exception:
        # Falls back to whatever was last entered in the Parâmetros sheet.
        # Surfaced in the UI as a warning so it's never silently stale.
        p = get_params()
        return (p['bank'] + p['cash'], p['bank'], p['cash'], 'fallback')

# ── Excel helpers ─────────────────────────────────────────────────────────────
def get_params():
    wb = openpyxl.load_workbook(NOVA_FILE, data_only=True)
    ws = wb['⚙️ Parâmetros']
    return {
        'bank': float(ws['B2'].value or 0),
        'cash': float(ws['B3'].value or 0),
        'loan1': float(ws['B5'].value or 100_000),
        'loan2': float(ws['B6'].value or 150_000),
        'da':   float(ws['B7'].value or 163_000),
    }

def load_gastos():
    wb = openpyxl.load_workbook(NOVA_FILE, data_only=True)
    ws = wb['💸 Gastos']
    rows, ynab_ids = [], set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        d, v = row[0], row[1]
        if v is None: continue
        dt = d.date() if isinstance(d, datetime) else (d if isinstance(d, date) else None)
        raw_h = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        ids = [x.strip() for x in raw_h.split(';') if x.strip()] if raw_h else []
        ynab_ids.update(ids)
        fatura = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        rows.append({
            'rownum': i, 'date': dt, 'amount': float(v),
            'desc': str(row[2] or ''), 'payee': str(row[3] or ''),
            'tag': str(row[4] or ''), 'conta': str(row[5] or ''),
            'ynab_ids': ids, 'fatura': fatura,
        })
    return rows, ynab_ids

def get_tags():
    rows, _ = load_gastos()
    extra = {r['tag'] for r in rows if r['tag']}
    return sorted(set(BASE_TAGS) | extra)

def link_ynab_id(rownum, txn_id):
    wb = openpyxl.load_workbook(NOVA_FILE)
    ws = wb['💸 Gastos']
    cell = ws.cell(rownum, 8)
    existing = [x.strip() for x in str(cell.value or '').split(';') if x.strip()]
    if txn_id not in existing:
        existing.append(txn_id)
    cell.value = ';'.join(existing)
    wb.save(NOVA_FILE)

def regen_dashboard():
    # Synchronous on purpose: a detached Popen here let dashboard.py re-save the
    # workbook while the next append_row was already writing it — two writers on
    # the same xlsx corrupted the zip (2026-08-10 incident).
    subprocess.run([sys.executable, DASHBOARD_PY], timeout=60)

def append_row(dt, amount, desc, payee, tag, conta, ynab_id=None, fatura=None,
               regen=True):
    wb = openpyxl.load_workbook(NOVA_FILE)
    ws = wb['💸 Gastos']
    r = ws.max_row + 1
    ws.cell(r, 1, datetime.combine(dt, datetime.min.time()) if isinstance(dt, date) else dt)
    ws.cell(r, 1).number_format = 'DD/MM/YYYY'
    ws.cell(r, 2, amount)
    ws.cell(r, 2).number_format = '#,##0.00 €'
    ws.cell(r, 3, desc)
    ws.cell(r, 4, payee)
    ws.cell(r, 5, tag)
    ws.cell(r, 6, conta)
    ws.cell(r, 7, conta)
    if ynab_id:
        ws.cell(r, 8, ynab_id)
    if fatura:
        ws.cell(r, 10, fatura)
    wb.save(NOVA_FILE)
    if regen:
        regen_dashboard()

# ── Analytics computation ─────────────────────────────────────────────────────
def build_analytics():
    rows, _ = load_gastos()
    params   = get_params()
    available, bank, cash, balance_source = ynab_snapshot()

    tag_totals = defaultdict(float)
    monthly    = defaultdict(float)

    for r in rows:
        tag_totals[r['tag']] += r['amount']
        if r['date']:
            monthly[(r['date'].year, r['date'].month)] += r['amount']

    total = sum(r['amount'] for r in rows)

    # Budget vs actual table
    cat_status = load_category_status()  # {tag: 'closed'}; default open
    table = []
    for tag in sorted(BUDGET.keys(), key=lambda x: -BUDGET[x]):
        actual = tag_totals.get(tag, 0)
        bc, bs = budget_civa(tag), BUDGET[tag]
        pct = actual / bc if bc else 0
        user_closed = cat_status.get(tag) == 'closed'
        if   actual == 0:        status, status_cls = 'Por iniciar', 'muted'
        elif user_closed:        status, status_cls = 'Fechada', 'green'
        elif actual > bc:        status, status_cls = 'Derrapagem', 'red'
        elif pct >= 0.95:        status, status_cls = 'Quase esgotado', 'orange'
        else:                    status, status_cls = 'Em curso', 'blue'
        table.append({
            'tag': tag, 'actual': actual, 'budget_siva': bs, 'budget_civa': bc,
            'remaining': max(0, bc - actual), 'overrun': max(0, actual - bc),
            'pct': pct, 'status': status, 'status_cls': status_cls,
            'closed': user_closed, 'has_budget': True,
        })
    for tag in ['Taxas & Multas', 'Empréstimo']:
        actual = tag_totals.get(tag, 0)
        table.append({
            'tag': tag, 'actual': actual, 'budget_siva': 0, 'budget_civa': 0,
            'remaining': 0, 'overrun': 0, 'pct': 0,
            'status': 'Sem orçamento', 'status_cls': 'muted',
            'closed': False, 'has_budget': False,
        })

    # Per-category progress (sorted by budget desc, for the bars below)
    categories = []
    for row in table:
        if row['budget_civa'] > 0:
            categories.append({
                'name': row['tag'], 'spent': row['actual'],
                'budget': row['budget_civa'],
                'pct': row['actual'] / row['budget_civa'] if row['budget_civa'] else 0,
            })
    categories.sort(key=lambda c: -c['budget'])

    # Spent-by-category chart (only categories with actuals, desc by amount)
    total_for_share = sum(r['actual'] for r in table if r['actual'] > 0)
    spent_by_cat = []
    for row in table:
        if row['actual'] > 0:
            spent_by_cat.append({
                'name': row['tag'], 'amount': round(row['actual'], 2),
                'pct': row['actual'] / total_for_share if total_for_share else 0,
            })
    spent_by_cat.sort(key=lambda c: -c['amount'])

    # Monthly chart data
    sorted_months = sorted(monthly.keys())
    monthly_labels = [f"{m:02d}/{y}" for (y, m) in sorted_months]
    monthly_values = [round(monthly[k], 2) for k in sorted_months]
    cum, cumulative = 0, []
    for v in monthly_values:
        cum += v; cumulative.append(round(cum, 2))

    # available comes from ynab_snapshot above — YNAB category 'available'

    # Projected c/IVA: per category — untouched → budget; closed → actual; else max(actual, budget)
    projected = 0
    for row in table:
        if not row['has_budget']:
            projected += row['actual']
        elif row['actual'] == 0:
            projected += row['budget_civa']
        elif row['closed']:
            projected += row['actual']
        else:
            projected += max(row['actual'], row['budget_civa'])

    # s/IVA version: reverse IVA on actuals (exempt stays)
    projected_siva = 0
    for row in table:
        actual_siva = row['actual'] if row['tag'] in IVA_EXEMPT \
                      else row['actual'] / (1 + IVA_RATE)
        if not row['has_budget']:
            projected_siva += actual_siva
        elif row['actual'] == 0:
            projected_siva += row['budget_siva']
        elif row['closed']:
            projected_siva += actual_siva
        else:
            projected_siva += max(actual_siva, row['budget_siva'])

    # Falta pagar: orçamento c/IVA ainda não consumido, excluindo fechadas
    remaining_to_pay = sum(
        max(0, row['budget_civa'] - row['actual'])
        for row in table if not row['closed']
    )

    # Recent entries (last 15)
    recent = sorted([r for r in rows if r['date']], key=lambda x: x['date'], reverse=True)[:15]

    return dict(
        total=total, budget_civa=TOTAL_CIVA, budget_siva=TOTAL_SIVA,
        pct=total / TOTAL_CIVA if TOTAL_CIVA else 0,
        available=available, bank=bank, cash=cash, balance_source=balance_source,
        projected=projected, projected_siva=projected_siva,
        remaining_to_pay=remaining_to_pay,
        da=params['da'],
        table=table, categories=categories, spent_by_cat=spent_by_cat, recent=recent,
        monthly_labels=json.dumps(monthly_labels),
        monthly_values=json.dumps(monthly_values),
        monthly_cumulative=json.dumps(cumulative),
    )

# ── Routes ────────────────────────────────────────────────────────────────────
@app.context_processor
def inject_demo_flag():
    """Make {{ demo }} available in all templates. Activated via ?demo=1."""
    return {'demo': request.args.get('demo') == '1'}

@app.route('/')
def index():
    return render_template('analytics.html', **build_analytics())

@app.route('/category/toggle', methods=['POST'])
def category_toggle():
    tag = request.form.get('tag', '').strip()
    if not tag: return redirect(url_for('index'))
    s = load_category_status()
    s[tag] = 'open' if s.get(tag) == 'closed' else 'closed'
    save_category_status(s)
    return redirect(url_for('index'))

@app.route('/add', methods=['GET', 'POST'])
def add():
    if request.method == 'POST':
        try:
            dt     = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
            amount = float(request.form['amount'].replace(',', '.'))
            desc   = request.form['desc'].strip()
            payee  = request.form['payee'].strip()
            new_tag = request.form.get('newtag', '').strip()
            tag    = new_tag or request.form['tag']
            conta  = request.form.get('conta', 'Cash')
            fatura = request.form.get('fatura', '').strip() or None

            account_id = CASH_OBRA_ID if conta == 'Cash' else MG_ID
            ynab_id = None
            try:
                resp = ynab_post(f'/budgets/{BUDGET_ID}/transactions', {'transaction': {
                    'account_id': account_id,
                    'date': dt.strftime('%Y-%m-%d'),
                    'amount': -int(amount * 1000),
                    'payee_name': payee or desc,
                    'memo': desc,
                    'category_id': CAT_ID,
                    'cleared': 'cleared',
                }})
                ynab_id = resp['data']['transaction']['id']
            except Exception as e:
                flash(f'Aviso YNAB: {e}', 'warning')

            append_row(dt, amount, desc, payee, tag, conta, ynab_id, fatura)
            _balance_cache['ts'] = 0.0  # force refresh — balance just changed
            flash(f'✓ Adicionado: €{amount:,.2f} — {desc}', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Erro: {e}', 'error')

    return render_template('add.html', tags=get_tags(), today=date.today().strftime('%Y-%m-%d'))

@app.route('/open-excel', methods=['POST'])
def open_excel():
    try:
        subprocess.Popen(['open', NOVA_FILE])
        return ('', 204)
    except Exception as e:
        return (str(e), 500)

@app.route('/sync')
def sync():
    try:
        rows, existing_ids = load_gastos()
        data  = ynab_get(f'/budgets/{BUDGET_ID}/categories/{CAT_ID}/transactions')
        txns  = data['data']['transactions']
        new   = [t for t in txns
                 if not t['deleted'] and t['amount'] < 0 and t['id'] not in existing_ids]
        new.sort(key=lambda t: t['date'], reverse=True)

        sorted_rows = sorted(rows, key=lambda r: (r['date'] or date.min), reverse=True)
        rows_dto = []
        for r in sorted_rows:
            d_str = r['date'].isoformat() if r['date'] else '—'
            parts = [d_str, r['payee'], r['desc'], r['tag'], r['conta'],
                     f"€{r['amount']:,.2f}"]
            parts = [p for p in parts if p]
            prefix = '🔗 ' if r['ynab_ids'] else ''
            rows_dto.append({'rownum': r['rownum'], 'label': prefix + ' · '.join(parts)})

        return render_template('sync.html',
            txns=new, tags=get_tags(), all_rows=rows_dto,
            txns_json=json.dumps(new))
    except Exception as e:
        flash(f'Erro ao contactar YNAB: {e}', 'error')
        return redirect(url_for('index'))

@app.route('/sync/confirm', methods=['POST'])
def sync_confirm():
    txns    = json.loads(request.form.get('txns_json', '[]'))
    txn_map = {t['id']: t for t in txns}
    imported = linked = ignored = 0

    for tid, t in txn_map.items():
        act = request.form.get(f'act_{tid}', 'ignore')

        if act == 'import':
            dt     = datetime.strptime(t['date'], '%Y-%m-%d').date()
            amount = abs(t['amount']) / 1000
            conta  = 'Cash' if t['account_id'] == CASH_OBRA_ID else 'Banco'
            new_tag = request.form.get(f'newtag_{tid}', '').strip()
            tag     = new_tag or request.form.get(f'tag_{tid}', '')
            desc    = t.get('memo') or t.get('payee_name', '')
            fatura  = request.form.get(f'fatura_{tid}', '').strip() or None
            append_row(dt, amount, desc, t.get('payee_name', ''), tag, conta, t['id'], fatura,
                       regen=False)  # one regen for the whole batch, below
            imported += 1

        elif act == 'link':
            rownum = request.form.get(f'link_{tid}', '').strip()
            if rownum.isdigit():
                link_ynab_id(int(rownum), t['id'])
                linked += 1

        else:
            ignored += 1

    if imported:
        regen_dashboard()
    if imported or linked:
        _balance_cache['ts'] = 0.0  # force refresh — YNAB state just changed

    msg = []
    if imported: msg.append(f'{imported} importada(s)')
    if linked:   msg.append(f'{linked} ligada(s) a linha existente')
    if ignored:  msg.append(f'{ignored} ignorada(s)')
    flash('✓ ' + ' · '.join(msg) if msg else 'Nada processado.', 'success')
    return redirect(url_for('index'))

# ── Assistente (Claude API) ───────────────────────────────────────────────────
# Embedded consultant chat. Claude gets the full budget/spending picture as
# cached context (prompt caching: persona + data snapshot form the stable
# prefix; the per-turn conversation comes after the cache breakpoint).
# The API key is loaded lazily so a missing key never crashes app startup
# (lesson learned from the YNAB token FileNotFoundError incident).

# Sonnet 5: near-Opus analysis quality at ~40% of the price (intro $2/$10 per
# MTok through 2026-08), and the cheapest model that supports the modern
# web_search tool (Haiku doesn't). effort=medium trims thinking spend further.
ASSISTANT_MODEL = 'claude-sonnet-5'

ASSISTANT_TOOLS = [{
    'type': 'web_search_20260209',
    'name': 'web_search',
    'max_uses': 3,
}]

ASSISTANT_PERSONA = """És um consultor sénior especializado em construção de \
moradias unifamiliares em Portugal, contratado pelo Luís como assessor pessoal \
para o projecto "Casa Nova" — a construção da casa da família dele.

## O teu papel
- Consultor técnico: fases de obra, sequência de trabalhos, boas práticas, \
materiais, normas portuguesas (RGEU, REH, licenciamento camarário, telas finais, \
livro de obra), IVA em obras (taxa normal 23%; autoliquidação quando aplicável).
- Analista financeiro do projecto: acompanhas o orçamento, detectas derrapagens, \
projectas custos até ao fim da obra, e ajudas a decidir onde apertar ou reforçar.
- Assessor pragmático: o Luís é gestor experiente (Head of Delivery Management \
em SaaS) — fala com ele de igual para igual, sem simplificações desnecessárias.

## Contexto do projecto
- Moradia unifamiliar em construção, financiada por: empréstimo bancário \
(linha de crédito aprovada, tranches libertadas por fases), contribuição \
familiar ("Dona Ana"), e capital próprio.
- A gestão financeira corre em YNAB (You Need A Budget) + folha Excel própria. \
Os dados que recebes abaixo vêm directamente dessas fontes.
- Categorias de obra usadas: Terreno, Arquitectura, Construção grosso, Pladur, \
Pichelaria & Climatização, Aluminio, Carpintaria, Construção acabamentos, \
Capoto, Electricidade, Pintura, Piscina, Estores, Furo, Design & Decoração \
Interiores, Serralheiro, Loiças & Cerâmicos, Electrodomésticos, Agua & Luz, \
Outros, Taxas & Multas, Empréstimo.
- Orçamentos por categoria estão definidos s/IVA; a comparação com gastos reais \
é feita c/IVA (23%), excepto Terreno (isento).

## Como responder
- Responde sempre em português de Portugal (norma europeia, pré-acordo: \
"projecto", "actual", "óptimo").
- Sê directo e concreto. Números sempre que possível. Não valides por validar.
- Quando fizeres análises, mostra o raciocínio de cálculo de forma compacta.
- Se detectares um risco (derrapagem, categoria suborçamentada, falta de \
liquidez prevista), di-lo proactivamente mesmo que não tenha sido perguntado.
- Se te faltar informação que os dados não cobrem, di-lo explicitamente — \
não inventes valores.
- Tens acesso a pesquisa na web. Usa-a quando a resposta depender de \
informação externa actual: preços de mercado de materiais/serviços, alterações \
a normas ou taxas, fornecedores, prazos típicos. Não pesquises para o que já \
está nos dados do projecto.
- Usa formatação markdown ligeira: negrito para números-chave, tabelas quando \
comparares categorias, listas curtas. Nada de headers pomposos em respostas curtas.
"""

_anthropic_client = None

def get_anthropic_client():
    """Lazy-init the Anthropic client. Key from env or key file."""
    global _anthropic_client
    if _anthropic_client is not None:
        return _anthropic_client
    import anthropic
    key = os.environ.get('ANTHROPIC_API_KEY')
    if not key:
        # Accept the plain .txt or the .rtf TextEdit tends to produce
        for path in (ANTHROPIC_KEY_FILE, ANTHROPIC_KEY_FILE + '.rtf',
                     ANTHROPIC_KEY_FILE.replace('.txt', '.rtf')):
            if os.path.exists(path):
                with open(path, errors='ignore') as f:
                    m = re.search(r'sk-ant-[A-Za-z0-9_-]{20,}', f.read())
                if m:
                    key = m.group(0)
                    break
    if not key:
        raise RuntimeError(
            f'Chave API Anthropic não encontrada. Cria o ficheiro '
            f'"{ANTHROPIC_KEY_FILE}" com a tua chave (sk-ant-...), '
            f'ou define ANTHROPIC_API_KEY.'
        )
    _anthropic_client = anthropic.Anthropic(api_key=key)
    return _anthropic_client

def build_assistant_snapshot():
    """Deterministic serialization of the full project state for the system
    prompt. No timestamps or volatile IDs — byte-stable between requests so
    the prompt-cache prefix survives across turns. Only changes when the
    underlying data changes (new expense, YNAB movement)."""
    a = build_analytics()
    rows, _ = load_gastos()
    params = get_params()

    lines = ['# Estado actual do projecto Casa Nova', '']
    lines.append('## Situação financeira')
    lines.append(f'- Disponibilidade Casa Nova (categoria YNAB): €{a["available"]:,.2f}')
    lines.append(f'- Repartição: Banco €{a["bank"]:,.2f} · Cash Obra €{a["cash"]:,.2f}'
                 f' (fonte: {a["balance_source"]})')
    lines.append(f'- Total gasto até hoje: €{a["total"]:,.2f}'
                 f' ({a["pct"]*100:.1f}% do orçamento c/IVA)')
    lines.append(f'- Orçamento total: €{a["budget_siva"]:,.2f} s/IVA'
                 f' = €{a["budget_civa"]:,.2f} c/IVA')
    lines.append(f'- Custo projectado final: €{a["projected"]:,.2f} c/IVA'
                 f' (€{a["projected_siva"]:,.2f} s/IVA)')
    lines.append(f'- Falta pagar (orçamento não consumido, excl. fechadas):'
                 f' €{a["remaining_to_pay"]:,.2f}')
    lines.append('')
    lines.append('## Financiamento')
    lines.append(f'- Empréstimo aprovado (linha de crédito): €310,000.00')
    lines.append(f'- Tranches libertadas: €{params["loan1"]:,.2f} (2025-02-21)'
                 f' + €{params["loan2"]:,.2f} (2026-05-12) = €{params["loan1"]+params["loan2"]:,.2f}')
    lines.append(f'- Margem de crédito ainda disponível: €{310000-params["loan1"]-params["loan2"]:,.2f}')
    lines.append(f'- Contribuição Dona Ana: €{params["da"]:,.2f} (totalmente entregue)')
    lines.append('')
    lines.append('## Orçamento vs real por categoria (valores c/IVA excepto indicação)')
    lines.append('| Categoria | Gasto | Orçamento c/IVA | % | Estado |')
    lines.append('|---|---|---|---|---|')
    for row in a['table']:
        if row['has_budget']:
            lines.append(f'| {row["tag"]} | €{row["actual"]:,.2f} |'
                         f' €{row["budget_civa"]:,.2f} | {row["pct"]*100:.0f}% |'
                         f' {row["status"]} |')
        else:
            lines.append(f'| {row["tag"]} | €{row["actual"]:,.2f} | — | — | sem orçamento |')
    lines.append('')
    lines.append('## Registo completo de gastos (data | valor | descrição | fornecedor | categoria | conta)')
    for r in sorted(rows, key=lambda x: (x['date'] or date.min)):
        d = r['date'].date().isoformat() if hasattr(r['date'], 'date') else (
            r['date'].isoformat() if r['date'] else '—')
        lines.append(f'{d} | €{r["amount"]:,.2f} | {r["desc"]} | {r["payee"]}'
                     f' | {r["tag"]} | {r["conta"]}')
    return '\n'.join(lines)

@app.route('/assistente')
def assistant():
    # The chat now lives in a drawer on every page (layout.html)
    return redirect(url_for('index'))

@app.route('/assistente/stream', methods=['POST'])
def assistant_stream():
    import anthropic

    payload  = request.get_json(force=True)
    messages = payload.get('messages', [])
    if not messages:
        return jsonify({'error': 'sem mensagens'}), 400

    try:
        client   = get_anthropic_client()
        snapshot = build_assistant_snapshot()
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 503
    except Exception as e:
        return jsonify({'error': f'Erro ao preparar contexto: {e}'}), 500

    system = [
        {'type': 'text', 'text': ASSISTANT_PERSONA},
        # Breakpoint on the snapshot: persona + data form the cached prefix.
        # Multi-turn conversations pay the cache-write once, then read.
        {'type': 'text', 'text': snapshot,
         'cache_control': {'type': 'ephemeral'}},
        # Volatile content after the breakpoint — never invalidates the cache.
        {'type': 'text', 'text': f'Data de hoje: {date.today().isoformat()}'},
    ]

    def generate():
        try:
            convo = list(messages)
            for _ in range(5):  # pause_turn continuation cap
                with client.messages.stream(
                    model=ASSISTANT_MODEL,
                    max_tokens=16000,
                    thinking={'type': 'adaptive'},
                    output_config={'effort': 'medium'},
                    system=system,
                    tools=ASSISTANT_TOOLS,
                    messages=convo,
                ) as stream:
                    for event in stream:
                        if (event.type == 'content_block_start'
                                and event.content_block.type == 'server_tool_use'):
                            yield 'data: ' + json.dumps(
                                {'status': 'a pesquisar na web…'}) + '\n\n'
                        elif (event.type == 'content_block_delta'
                                and event.delta.type == 'text_delta'):
                            yield 'data: ' + json.dumps(
                                {'text': event.delta.text}, ensure_ascii=False) + '\n\n'
                    final = stream.get_final_message()
                # Server-side tool loop hit its iteration limit — resume where
                # it left off by echoing the assistant turn back.
                if final.stop_reason == 'pause_turn':
                    convo = convo + [{'role': 'assistant', 'content': final.content}]
                    continue
                break
            yield 'data: ' + json.dumps({
                'done': True,
                'cache_read': final.usage.cache_read_input_tokens,
                'cache_write': final.usage.cache_creation_input_tokens,
            }) + '\n\n'
        except anthropic.AuthenticationError:
            yield 'data: ' + json.dumps(
                {'error': 'Chave API inválida — verifica o ficheiro da chave.'}) + '\n\n'
        except anthropic.RateLimitError:
            yield 'data: ' + json.dumps(
                {'error': 'Rate limit atingido — tenta de novo daqui a um minuto.'}) + '\n\n'
        except anthropic.APIStatusError as e:
            yield 'data: ' + json.dumps(
                {'error': f'Erro da API ({e.status_code}): {e.message}'}) + '\n\n'
        except anthropic.APIConnectionError:
            yield 'data: ' + json.dumps(
                {'error': 'Sem ligação à API Anthropic — verifica a internet.'}) + '\n\n'

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

# ── Launch ────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception:
        local_ip = '127.0.0.1'

    iphone_url = f'http://{local_ip}:5001'
    print()
    print('━' * 52)
    print('  🏗  Casa Nova')
    print(f'  Mac:     http://localhost:5001')
    print(f'  iPhone:  {iphone_url}')
    print('━' * 52)

    try:
        qr = qrcode.QRCode(border=1)
        qr.add_data(iphone_url)
        qr.make(fit=True)
        qr.print_ascii(invert=True)
    except Exception:
        pass

    print()
    subprocess.Popen(['open', 'http://localhost:5001'])
    app.run(host='0.0.0.0', port=5001, debug=False)
