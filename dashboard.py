#!/usr/bin/env python3
"""
Casa Nova — Dashboard Builder (formula-based)
All metrics use Excel SUMIF/SUMPRODUCT formulas referencing '💸 Gastos'.
A '⚙️ Parâmetros' sheet holds editable reference values (YNAB balances, loan info).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import date
from collections import defaultdict

NOVA_FILE = '/Users/luisfbarbosa/Documents/Claude/Casa Nova/Custos Casa Nova - Nova2.xlsx'
G         = "'💸 Gastos'"      # sheet reference for formulas
MAX_ROW   = 500                # formula range ceiling

# Current YNAB balances (written to Params sheet — user updates manually)
YNAB_BANK   = 24466.27
YNAB_CASH   = 16700.00
LOAN_DATE1  = date(2025, 2, 21)
LOAN_AMT1   = 100_000
LOAN_AMT2   = 150_000          # planned
DA_TOTAL    = 163_000

# IVA settings
IVA_RATE    = 0.23          # standard Portuguese IVA for construction
IVA_EXEMPT  = {'Terreno'}   # land purchase — no IVA

# Budget s/IVA by tag (agreed contract values)
BUDGET = {
    'Terreno':                       70_000,
    'Arquitectura':                   4_500,
    'Construção grosso':             80_000,
    'Pladur':                        22_000,
    'Pichelaria & Climatização':     25_000,
    'Aluminio':                      25_000,
    'Carpintaria':                   40_000,
    'Construção acabamentos':        30_000,
    'Capoto':                        15_000,
    'Electricidade':                 15_000,
    'Pintura':                       15_000,
    'Piscina':                       18_000,
    'Estores':                        7_000,
    'Furo':                           5_000,
    'Design & Decoração Interiores': 10_400,
    'Serralheiro':                   10_000,
    'Loiças & Cerâmicos':             8_000,
    'Electrodomésticos':             10_000,
    'Agua & Luz':                     1_000,
    'Outros':                        10_000,
}

def budget_civa(tag):
    """Budget with IVA for a given tag."""
    b = BUDGET.get(tag, 0)
    return b if tag in IVA_EXEMPT else round(b * (1 + IVA_RATE), 2)

TOTAL_BUDGET_SIVA = sum(BUDGET.values())
TOTAL_BUDGET      = sum(budget_civa(t) for t in BUDGET)  # c/IVA — used for planning

PHASES = [
    ('Terreno & Licenças',     ['Terreno', 'Taxas & Multas', 'Arquitectura', 'Empréstimo']),
    ('Estrutura',              ['Construção grosso', 'Furo']),
    ('Envolvente exterior',    ['Capoto', 'Aluminio', 'Serralheiro']),
    ('Instalações',            ['Pichelaria & Climatização', 'Electricidade']),
    ('Acabamentos interiores', ['Construção acabamentos', 'Pladur', 'Pintura', 'Carpintaria']),
    ('Equipamentos & Deco',    ['Electrodomésticos', 'Design & Decoração Interiores',
                                'Loiças & Cerâmicos', 'Estores']),
    ('Infra & Outros',         ['Agua & Luz', 'Outros', 'Piscina']),
]

# ── Styles ────────────────────────────────────────────────────────────────────
BLUE_DARK  = '1F497D'
BLUE_MID   = '4472C4'
BLUE_LIGHT = 'D9E1F2'
GREEN_F    = 'E2EFDA'
RED_F      = 'FFE2E2'
ORANGE_F   = 'FFF2CC'
GREY_F     = 'F2F2F2'
WHITE      = 'FFFFFF'
EUR        = '#,##0.00 €'
PCT        = '0.0"%"'

def hf(c):  return PatternFill('solid', fgColor=c)
def bf(bold=True, size=10, color='000000'): return Font(bold=bold, size=size, color=color)
def af(h='left'): return Alignment(horizontal=h, vertical='center')

def header(ws, r, c1, c2, text, bg=BLUE_MID, fg=WHITE, size=11, height=20):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.font      = Font(bold=True, size=size, color=fg)
    cell.fill      = hf(bg)
    cell.alignment = af('center')
    ws.row_dimensions[r].height = height
    return cell

def kpi(ws, r, c, label, value_or_formula, fmt=EUR, lbg=BLUE_LIGHT, vbg=WHITE):
    lc = ws.cell(r, c, label)
    lc.font, lc.fill, lc.alignment = bf(bold=True), hf(lbg), af('left')
    vc = ws.cell(r, c + 1, value_or_formula)
    vc.fill, vc.alignment = hf(vbg), af('right')
    if fmt: vc.number_format = fmt
    return vc

def sumif_tag(tag):
    return f'=SUMIF({G}!E:E,"{tag}",{G}!B:B)'

def sumproduct_tag_today(tag):
    return (f'=SUMPRODUCT(ISNUMBER({G}!A2:A{MAX_ROW})'
            f'*({G}!A2:A{MAX_ROW}<=TODAY())'
            f'*({G}!E2:E{MAX_ROW}="{tag}")'
            f'*({G}!B2:B{MAX_ROW}))')

def sumproduct_conta_since(conta, since_cell):
    """Sum of entries for a given Conta value on or after a date in since_cell."""
    return (f'=SUMPRODUCT(ISNUMBER({G}!A2:A{MAX_ROW})'
            f'*({G}!A2:A{MAX_ROW}>={since_cell})'
            f'*({G}!G2:G{MAX_ROW}="{conta}")'
            f'*({G}!B2:B{MAX_ROW}))')

# ── Build Parâmetros sheet ────────────────────────────────────────────────────

def build_params(wb):
    name = '⚙️ Parâmetros'
    if name in wb.sheetnames: del wb[name]
    ws = wb.create_sheet(name)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    header(ws, 1, 1, 2, '⚙️ Parâmetros — actualiza manualmente', BLUE_DARK, WHITE, 11)

    rows = [
        ('YNAB — Banco (MG Ordenado)',    YNAB_BANK,              EUR,          BLUE_LIGHT),
        ('YNAB — Cash (Cache obra)',       YNAB_CASH,              EUR,          BLUE_LIGHT),
        ('Empréstimo 1ª tranche — data',  LOAN_DATE1,             'DD/MM/YYYY', BLUE_LIGHT),
        ('Empréstimo 1ª tranche — valor', LOAN_AMT1,              EUR,          BLUE_LIGHT),
        ('Empréstimo 2ª tranche — valor', LOAN_AMT2,              EUR,          BLUE_LIGHT),
        ('Dona Ana — contribuição total', DA_TOTAL,               EUR,          BLUE_LIGHT),
        ('Taxa IVA obras (ajusta se necessário)', IVA_RATE,       '0%',         ORANGE_F),
        ('Orçamento total s/IVA',         TOTAL_BUDGET_SIVA,      EUR,          GREY_F),
        ('Orçamento total c/IVA',         TOTAL_BUDGET,           EUR,          GREY_F),
    ]
    for i, (label, val, fmt, bg) in enumerate(rows, 2):
        kpi(ws, i, 1, label, val, fmt, bg)

    ws.cell(12, 1, 'Nota: actualiza B2 e B3 após cada extracto YNAB. B8 = taxa IVA aplicada ao orçamento.')
    ws.cell(12, 1).font = Font(italic=True, size=9, color='808080')

    return ws, {
        'bank':        "'⚙️ Parâmetros'!B2",
        'cash':        "'⚙️ Parâmetros'!B3",
        'loan_d':      "'⚙️ Parâmetros'!B4",
        'loan1':       "'⚙️ Parâmetros'!B5",
        'loan2':       "'⚙️ Parâmetros'!B6",
        'da':          "'⚙️ Parâmetros'!B7",
        'iva_rate':    "'⚙️ Parâmetros'!B8",
        'total_civa':  "'⚙️ Parâmetros'!B10",
    }

# ── Build Dashboard sheet ─────────────────────────────────────────────────────

def build_dashboard(wb, p):
    name = '📊 Dashboard'
    if name in wb.sheetnames: del wb[name]
    ws = wb.create_sheet(name, 1)

    for c, w in {1:34, 2:16, 3:3, 4:34, 5:16, 6:3, 7:14, 8:14, 9:14, 10:10, 11:14}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

    r = 1
    header(ws, r, 1, 11, '🏗  CASA NOVA — PAINEL DE CONTROLO FINANCEIRO',
           BLUE_DARK, WHITE, 14, 28)
    r += 1
    ws.cell(r, 1, f'Fórmulas ligadas a "💸 Gastos". Actualiza os saldos YNAB em "⚙️ Parâmetros".')
    ws.cell(r, 1).font = Font(italic=True, size=9, color='808080')
    r += 2

    # ── 1. Project snapshot ───────────────────────────────────────────────────
    header(ws, r, 1, 5, '📍 Resumo do Projecto', BLUE_MID)
    header(ws, r, 7, 11, '💰 Liquidez Actual (YNAB)', BLUE_MID)
    r += 1

    kpi(ws, r, 1, 'Total gasto (todas as entradas)',
        f'=SUM({G}!B:B)')
    kpi(ws, r, 7, 'Banco — MG Ordenado',
        f'={p["bank"]}')
    r += 1
    kpi(ws, r, 1, 'Total gasto até hoje',
        f'=SUMPRODUCT(ISNUMBER({G}!A2:A{MAX_ROW})*({G}!A2:A{MAX_ROW}<=TODAY())*({G}!B2:B{MAX_ROW}))')
    kpi(ws, r, 7, 'Cash — Cache obra',
        f'={p["cash"]}')
    r += 1
    kpi(ws, r, 1, 'Orçamento total c/IVA',
        TOTAL_BUDGET)
    kpi(ws, r, 7, 'Total disponível',
        f'={p["bank"]}+{p["cash"]}',
        vbg=GREEN_F)
    r += 1
    kpi(ws, r, 1, '% orçamento executado (vs. c/IVA)',
        f'=SUM({G}!B:B)/{TOTAL_BUDGET}', PCT)
    kpi(ws, r, 7, 'Empréstimo MG — 1ª tranche',
        f'={p["loan1"]}')
    r += 1
    # Projected final and overrun — formulas filled in after budget table (rows unknown here)
    kpi(ws, r, 1, 'Custo projectado conclusão', '', EUR, ORANGE_F)
    projected_row = r
    kpi(ws, r, 7, 'Empréstimo MG — 2ª tranche (prevista)',
        f'={p["loan2"]}')
    r += 1
    kpi(ws, r, 1, 'Derrapagem (categorias acima do orçamento)', '', EUR, RED_F)
    overrun_kpi_row = r
    # Défice = remaining to spend (projected - spent) minus available cash today
    # Projected total includes already-spent, so subtract it to get what's still to pay
    kpi(ws, r, 7, 'Défice antes de 2ª tranche (falta pagar vs. liquidez actual)',
        f'=MAX(0,B{projected_row}-SUM({G}!B:B)-{p["bank"]}-{p["cash"]})',
        vbg=RED_F)
    r += 2

    # ── 2. Budget vs Actual ───────────────────────────────────────────────────
    header(ws, r, 1, 11, '📊 Orçamento vs Real — por Categoria', BLUE_MID)
    r += 1

    # Columns: Cat(1) | Actual(2) | sp(3) | Budget s/IVA(4) | Budget c/IVA(5) | A Pagar(6) | sp(7) | Derrapagem(8) | %(9) | Estado(10)
    col_headers = ['Categoria', 'Pago (real c/IVA)', 'Orçamento s/IVA', 'Orçamento c/IVA', 'A Pagar c/IVA', 'Derrapagem', '% Exec.', 'Estado']
    col_cols    = [1,            2,                   4,                  5,                  6,               8,            9,          10]
    for h, c in zip(col_headers, col_cols):
        cell = ws.cell(r, c, h)
        cell.font, cell.fill = bf(bold=True, color=WHITE), hf(BLUE_DARK)
        cell.alignment = af('center')
    ws.row_dimensions[r].height = 18
    r += 1

    budget_table_start = r
    all_tags = sorted(set(list(BUDGET.keys()) + ['Taxas & Multas', 'Empréstimo']),
                      key=lambda x: BUDGET.get(x, 0), reverse=True)

    for tag in all_tags:
        bgt_siva  = BUDGET.get(tag, 0)
        bgt_civa  = budget_civa(tag)
        actual_cell     = f'B{r}'
        budget_civa_cell = f'E{r}'   # column 5 = E
        ws.cell(r, 1, tag).font = Font(size=10)

        ws.cell(r, 2, sumif_tag(tag)).number_format = EUR
        ws.cell(r, 2).font = Font(size=10)

        ws.cell(r, 4, bgt_siva if bgt_siva else None)
        ws.cell(r, 4).number_format = EUR
        ws.cell(r, 4).font = Font(size=10)

        ws.cell(r, 5, bgt_civa if bgt_civa else None)
        ws.cell(r, 5).number_format = EUR
        ws.cell(r, 5).font = Font(size=10)

        ws.cell(r, 6, f'=MAX(0,{budget_civa_cell}-{actual_cell})').number_format = EUR
        ws.cell(r, 6).font = Font(size=10)

        ws.cell(r, 8, f'=MAX(0,{actual_cell}-{budget_civa_cell})').number_format = EUR
        ws.cell(r, 8).font = Font(size=10)

        pct_f = f'=IF({budget_civa_cell}>0,{actual_cell}/{budget_civa_cell},"-")'
        ws.cell(r, 9, pct_f)
        ws.cell(r, 9).number_format = PCT
        ws.cell(r, 9).font = Font(size=10)

        ws.cell(r, 10, (
            f'=IF({budget_civa_cell}=0,"⚠ sem orçamento",'
            f'IF({actual_cell}>{budget_civa_cell},"🔴 Derrapagem",'
            f'IF({actual_cell}/{budget_civa_cell}>=0.95,"🟡 Quase esgotado",'
            f'IF({actual_cell}=0,"⬜ Não iniciado",'
            f'"🔵 Em curso"))))'
        ))
        ws.cell(r, 10).font = Font(size=10)
        r += 1

    budget_table_end = r - 1

    # Back-fill the KPI cells that need budget table row references
    # Projected: actuals + full c/IVA budget for unstarted categories (col E = budget c/IVA, col B = actual)
    ws.cell(projected_row, 2).value = (
        f'=SUM({G}!B:B)+SUMPRODUCT('
        f'(E{budget_table_start}:E{budget_table_end}>0)'
        f'*(B{budget_table_start}:B{budget_table_end}=0)'
        f'*E{budget_table_start}:E{budget_table_end})'
    )
    ws.cell(projected_row, 2).number_format = EUR
    ws.cell(overrun_kpi_row, 2).value = (
        f'=SUMPRODUCT('
        f'(B{budget_table_start}:B{budget_table_end}>E{budget_table_start}:E{budget_table_end})'
        f'*(E{budget_table_start}:E{budget_table_end}>0)'
        f'*(B{budget_table_start}:B{budget_table_end}-E{budget_table_start}:E{budget_table_end})'
        f')'
    )
    ws.cell(overrun_kpi_row, 2).number_format = EUR

    # Totals
    for c, v, fmt in [
        (1,  'TOTAL', None),
        (2,  f'=SUM(B{budget_table_start}:B{budget_table_end})', EUR),
        (4,  TOTAL_BUDGET_SIVA, EUR),
        (5,  TOTAL_BUDGET, EUR),
        (6,  f'=SUM(F{budget_table_start}:F{budget_table_end})', EUR),
        (8,  f'=SUM(H{budget_table_start}:H{budget_table_end})', EUR),
    ]:
        cell = ws.cell(r, c, v)
        cell.font = bf(bold=True)
        cell.fill = hf(BLUE_LIGHT)
        cell.alignment = af('right' if c != 1 else 'left')
        if fmt: cell.number_format = fmt
    r += 2

    # ── 3. Construction phases ────────────────────────────────────────────────
    header(ws, r, 1, 5, '🏗 Fases da Construção', BLUE_MID)
    r += 1
    for h, c in zip(['Fase', 'Gasto', '% do Total'], [1, 2, 4]):
        ws.cell(r, c, h).font = bf(bold=True, color=WHITE)
        ws.cell(r, c).fill = hf(BLUE_DARK)

    r += 1
    phase_total_cells = []
    for phase_name, tags in PHASES:
        sumif_parts = '+'.join(f'SUMIF({G}!E:E,"{t}",{G}!B:B)' for t in tags)
        ws.cell(r, 1, phase_name).font = Font(size=10)
        ws.cell(r, 2, f'={sumif_parts}').number_format = EUR
        ws.cell(r, 2).font = Font(size=10)
        phase_total_cells.append(f'B{r}')
        r += 1

    grand_total_ref = '+'.join(phase_total_cells)
    for i, cell_ref in enumerate(phase_total_cells):
        phase_r = budget_table_end + 6 + i  # approximate row - recompute properly
    # Re-scan to add % column
    phase_start_r = r - len(PHASES)
    for i in range(len(PHASES)):
        pr = phase_start_r + i
        ws.cell(pr, 4, f'=B{pr}/({"+".join(f"B{phase_start_r+j}" for j in range(len(PHASES)))})').number_format = PCT
        ws.cell(pr, 4).font = Font(size=10)
    r += 1

    # ── 4. Monthly spend ─────────────────────────────────────────────────────
    header(ws, r, 1, 5, '📅 Gastos Mensais', BLUE_MID)
    r += 1
    for h, c in zip(['Ano', 'Mês', 'Gasto', 'Acumulado'], [1, 2, 4, 5]):
        ws.cell(r, c, h).font = bf(bold=True, color=WHITE)
        ws.cell(r, c).fill = hf(BLUE_DARK)
    r += 1

    # Load actual monthly data to write as formula-friendly table
    wb_read = openpyxl.load_workbook(NOVA_FILE, data_only=True)
    ws_g    = wb_read['💸 Gastos']
    from collections import defaultdict
    from datetime import datetime
    monthly = defaultdict(float)
    for row in ws_g.iter_rows(min_row=2, values_only=True):
        d, v = row[0], row[1]
        if not isinstance(v, (int, float)): continue
        dt = d.date() if isinstance(d, datetime) else (d if isinstance(d, date) else None)
        if dt: monthly[(dt.year, dt.month)] += v

    month_names = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
                   7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}
    monthly_row_start = r
    cum = 0
    for (yr, mo) in sorted(monthly):
        amt = monthly[(yr, mo)]
        cum += amt
        ws.cell(r, 1, yr).font = Font(size=10)
        ws.cell(r, 2, month_names[mo]).font = Font(size=10)
        ws.cell(r, 4, amt)
        ws.cell(r, 4).number_format = EUR
        ws.cell(r, 4).font = Font(size=10)
        ws.cell(r, 5, f'=SUM(D{monthly_row_start}:D{r})')
        ws.cell(r, 5).number_format = EUR
        ws.cell(r, 5).font = Font(size=10)
        r += 1
    monthly_row_end = r - 1
    r += 1

    # ── 5. Funding sources ────────────────────────────────────────────────────
    header(ws, r, 1, 5, '🏦 Fontes de Financiamento', BLUE_MID)
    header(ws, r, 7, 11, '🏦 Conta — Banco vs Cash', BLUE_MID)
    r += 1

    kpi(ws, r, 1, 'Dona Ana — contribuição total (referência)',
        f'={p["da"]}')
    # Conta completion indicator
    conta_filled_f  = f'=COUNTIF({G}!G:G,"Banco")+COUNTIF({G}!G:G,"Cash")'
    conta_total_f   = f'=COUNTA({G}!B2:B{MAX_ROW})'
    kpi(ws, r, 7, 'Entradas com Conta preenchida',
        f'={conta_filled_f}&" / "&{conta_total_f}&" (" &TEXT({conta_filled_f}/{conta_total_f},"0%")&")"',
        fmt=None, vbg=ORANGE_F if True else WHITE)
    r += 1
    kpi(ws, r, 1, 'Empréstimo 1ª tranche (fev 2025)',
        f'={p["loan1"]}')
    kpi(ws, r, 7, 'Pago via Banco  ⚠ incompleto até Conta estar 100%',
        f'=SUMIF({G}!G:G,"Banco",{G}!B:B)')
    r += 1
    kpi(ws, r, 1, 'Empréstimo 2ª tranche (prevista)',
        f'={p["loan2"]}')
    kpi(ws, r, 7, 'Pago via Cash  ⚠ incompleto até Conta estar 100%',
        f'=SUMIF({G}!G:G,"Cash",{G}!B:B)')
    r += 1
    kpi(ws, r, 1, 'Total financiamento previsto (DA + 1ª + 2ª tranche)',
        f'={p["da"]}+{p["loan1"]}+{p["loan2"]}', vbg=BLUE_LIGHT)
    kpi(ws, r, 7, 'Sem Conta preenchida',
        f'=SUMIF({G}!G:G,"",{G}!B:B)', vbg=ORANGE_F)
    r += 1
    loan_consumed_f = (
        f'=MIN({p["loan1"]},'
        f'SUMPRODUCT(ISNUMBER({G}!A2:A{MAX_ROW})'
        f'*({G}!A2:A{MAX_ROW}>={p["loan_d"]})'
        f'*({G}!B2:B{MAX_ROW})))'
    )
    kpi(ws, r, 1, 'Empréstimo consumido (aprox. — toda a despesa desde fev 2025)',
        loan_consumed_f, vbg=ORANGE_F)
    r += 1
    kpi(ws, r, 1, 'Empréstimo restante 1ª tranche (aprox.)',
        f'=MAX(0,{p["loan1"]}-B{r-1})', vbg=GREEN_F)
    r += 2

    # ── Chart: Budget vs Actual ───────────────────────────────────────────────
    header(ws, r, 7, 11, '📈 Orçamento vs Pago — por Categoria', BLUE_MID)
    r += 1
    chart_data_row = r
    ws.cell(r, 7, 'Categoria').font = bf(bold=True, size=9)
    ws.cell(r, 8, 'Orçamento').font = bf(bold=True, size=9)
    ws.cell(r, 9, 'Pago').font = bf(bold=True, size=9)
    r += 1
    chart_body_start = r
    for tag in sorted(BUDGET.keys(), key=lambda x: -BUDGET[x]):
        ws.cell(r, 7, tag)
        ws.cell(r, 8, budget_civa(tag))  # c/IVA budget for chart
        ws.cell(r, 9, sumif_tag(tag)).number_format = EUR
        r += 1
    chart_body_end = r - 1

    bar = BarChart()
    bar.type, bar.grouping = 'bar', 'clustered'
    bar.title, bar.style, bar.height, bar.width = None, 10, 14, 20
    cats = Reference(ws, min_col=7, min_row=chart_body_start, max_row=chart_body_end)
    bar.add_data(Reference(ws, min_col=8, min_row=chart_data_row, max_row=chart_body_end), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=9, min_row=chart_data_row, max_row=chart_body_end), titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = BLUE_MID
    bar.series[1].graphicalProperties.solidFill = '70AD47'
    ws.add_chart(bar, f'G{chart_data_row - 1}')
    r += 2

    # ── Chart: Cumulative spend ───────────────────────────────────────────────
    header(ws, r, 7, 11, '📈 Investimento Acumulado', BLUE_MID)
    r += 1
    line_hdr = r
    ws.cell(r, 7, 'Período').font = bf(bold=True, size=9)
    ws.cell(r, 8, 'Acumulado').font = bf(bold=True, size=9)
    r += 1
    line_start = r
    for (yr, mo) in sorted(monthly):
        ws.cell(r, 7, f'{yr}-{mo:02d}')
        ws.cell(r, 8, f'=SUM(D{monthly_row_start}:D{monthly_row_start + list(sorted(monthly)).index((yr, mo))})')
        ws.cell(r, 8).number_format = EUR
        r += 1
    line_end = r - 1

    lc = LineChart()
    lc.title, lc.style, lc.height, lc.width, lc.legend = None, 10, 10, 20, None
    lc.add_data(Reference(ws, min_col=8, min_row=line_hdr, max_row=line_end), titles_from_data=True)
    lc.set_categories(Reference(ws, min_col=7, min_row=line_start, max_row=line_end))
    lc.series[0].graphicalProperties.line.solidFill = BLUE_DARK
    lc.series[0].graphicalProperties.line.width = 20000
    ws.add_chart(lc, f'G{line_hdr}')

    return ws

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    wb = openpyxl.load_workbook(NOVA_FILE)
    print('Building Parâmetros sheet...')
    _, p = build_params(wb)
    print('Building Dashboard sheet...')
    build_dashboard(wb, p)
    wb.save(NOVA_FILE)
    print(f'✓ Saved to {NOVA_FILE}')
    print()
    print('Key reference values written to ⚙️ Parâmetros:')
    print(f'  YNAB Banco:    €{YNAB_BANK:,.2f}  ← update after each YNAB sync')
    print(f'  YNAB Cash:     €{YNAB_CASH:,.2f}  ← update after each YNAB sync')
    print(f'  Loan tranche 1: €{LOAN_AMT1:,.0f} (fev 2025)')
    print(f'  Loan tranche 2: €{LOAN_AMT2:,.0f} (prevista)')
    print(f'  Dona Ana:       €{DA_TOTAL:,.0f}')
