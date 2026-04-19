#!/usr/bin/env python3
"""
Casa Nova — Reconciliation Script
Merges YNAB 'Gastos Casa Nova' transactions with the existing Excel spreadsheet.
Produces a clean new spreadsheet with all transactions merged, sorted, and tagged.

Yellow rows = YNAB bank transactions not previously in the spreadsheet (review these).
White rows  = existing entries (cash, pre-YNAB, or already matched).
"""

import os, re, requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from collections import defaultdict
from itertools import combinations

# --- Config ---
# Token is read from an RTF file outside the repo — never commit secrets.
TOKEN_FILE  = os.path.expanduser('~/Documents/Finance/ynab token.rtf')
def _load_token():
    with open(TOKEN_FILE) as f:
        m = re.search(r'(?<![A-Za-z0-9_-])[A-Za-z0-9_-]{43}(?![A-Za-z0-9_-])', f.read())
    if not m: raise RuntimeError(f'No YNAB token found in {TOKEN_FILE}')
    return m.group(0)
YNAB_TOKEN  = _load_token()
BUDGET_ID   = "62da99f2-0125-45cd-8906-6a5ebe3416ad"
CAT_ID      = "6bc4c34f-60e3-462b-9f75-a9baf10d35b0"  # Gastos Casa Nova
SOURCE      = "/Users/luisfbarbosa/Downloads/Custos Casa Nova.xlsx"
OUTPUT      = "/Users/luisfbarbosa/Downloads/Custos Casa Nova - Nova.xlsx"

# Source of truth: current YNAB balances
YNAB_BANK   = 24466.27   # MG Ordenado
YNAB_CASH   = 16700.00   # Cache obra (10k withdrawal + 6,700 DA)
YNAB_TOTAL  = YNAB_BANK + YNAB_CASH

# --- Helpers ---

def parse_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for f in ['%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y']:
            try: return datetime.strptime(v, f).date()
            except: pass
    return None

def eval_amount(v):
    if isinstance(v, (int, float)): return round(float(v), 2)
    if isinstance(v, str) and v.startswith('='):
        expr = v[1:]
        if re.match(r'^[\d\+\-\*\/\(\)\. ]+$', expr):
            try: return round(float(eval(expr)), 2)
            except: pass
    return None

# Payee → Tag inference (for YNAB-only entries)
PAYEE_TAG = [
    (['agere', 'dd-agere', 'edp comercial', 'dd-edp'],          'Agua & Luz'),
    (['xptobra', 'xpto'],                                         'Construção acabamentos'),
    (['e-redes', 'eredes', 'pag.serv. 23565'],                   'Taxas & Multas'),
    (['câmara', 'camara', 'balcao unico', 'balcão único',
      'registo civil', 'imposto', 'multa', 'taxa', 'at '],       'Taxas & Multas'),
    (['aquecitubo', 'climatiz', 'pichelaria'],                    'Pichelaria & Climatização'),
    (['marmores', 'mármores', 'lazaro', 'pedras soleiras'],      'Construção grosso'),
    (['mdl', 'aluminio', 'alumínio', 'marcio paredes'],          'Aluminio'),
    (['placomoreira', 'pladur'],                                  'Pladur'),
    (['hidrocávado', 'hidrocavado', 'furo'],                      'Furo'),
    (['plataforma legal', 'com.estud', 'com.avaliacao',
      'i.selo', 'vistoria'],                                     'Empréstimo'),
    (['narrativa', 'andreia', 'design', 'interior'],             'Design & Decoração Interiores'),
    (['capoto', 'fernando dias'],                                 'Capoto'),
    (['serralheiro', 'serralharia', 'pimenta', 'luís carlos'],   'Serralheiro'),
    (['electrodoméstico', 'cidadela'],                            'Electrodomésticos'),
    (['arquitect', 'arquitec'],                                   'Arquitectura'),
]

def infer_tag(payee, memo=''):
    text = f"{payee} {memo}".lower()
    for keywords, tag in PAYEE_TAG:
        if any(k in text for k in keywords):
            return tag
    return ''

# --- 1. Fetch YNAB transactions ---

def fetch_ynab():
    resp = requests.get(
        f"https://api.ynab.com/v1/budgets/{BUDGET_ID}/transactions",
        headers={"Authorization": f"Bearer {YNAB_TOKEN}"},
        params={"since_date": "2021-01-01"}
    ).json()
    return [t for t in resp['data']['transactions']
            if t['category_id'] == CAT_ID
            and not t['deleted']
            and t['amount'] < 0]  # outflows only — inflows (e.g. loan tranches) are not expenses

# --- 2. Load existing spreadsheet ---

def load_sheet():
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb['💸Gastos']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        d, v, desc, payee, tag, fonte = row[0], row[1], row[2], row[3], row[4], row[5]
        ref   = row[6] if len(row) > 6 else None
        drive = row[7] if len(row) > 7 else None
        if d is None and v is None:
            continue
        dt  = parse_date(d)
        amt = eval_amount(v)
        if dt is None or amt is None:
            continue
        rows.append({
            'date':      dt,
            'amount':    amt,
            'desc':      str(desc  or ''),
            'payee':     str(payee or ''),
            'tag':       str(tag   or ''),
            'fonte':     str(fonte or ''),
            'ref':       str(ref   or ''),
            'drive':     str(drive or ''),
            'ynab_id':   None,
            'ynab_only': False,
        })
    return rows

# --- 3. Match and merge ---

# Payees that are bank-to-cash transfers, not direct expenses
SKIP_PAYEES = ['levantamento numerario']

def date_variants(dt):
    """Return plausible date interpretations to handle common entry errors."""
    variants = {dt}
    # Year typo: 2015 → 2025
    if dt.year == 2015:
        try: variants.add(dt.replace(year=2025))
        except ValueError: pass
    # Day/month swap (only when both values are valid as either day or month)
    if dt.day <= 12 and dt.month != dt.day:
        try: variants.add(dt.replace(month=dt.day, day=dt.month))
        except ValueError: pass
    # Also apply swap to the year-corrected version
    for vd in list(variants):
        if vd.day <= 12 and vd.month != vd.day and vd != dt:
            try: variants.add(vd.replace(month=vd.day, day=vd.month))
            except ValueError: pass
    return variants

def match_and_merge(sheet_rows, ynab_txns):
    used_sheet = set()
    used_ynab  = set()
    ynab_sorted = sorted(range(len(ynab_txns)), key=lambda i: ynab_txns[i]['date'])

    def run_pass(tolerance_days):
        for yi in ynab_sorted:
            if yi in used_ynab: continue
            txn = ynab_txns[yi]
            yd  = datetime.strptime(txn['date'], '%Y-%m-%d').date()
            ya  = round(abs(txn['amount']) / 1000, 2)
            best, best_delta = None, timedelta(days=tolerance_days + 1)
            for si, row in enumerate(sheet_rows):
                if si in used_sheet: continue
                if abs(row['amount'] - ya) > 0.02: continue
                for vd in date_variants(row['date']):
                    delta = abs(vd - yd)
                    if delta <= timedelta(days=tolerance_days) and delta < best_delta:
                        best, best_delta = si, delta
            if best is not None:
                sheet_rows[best]['ynab_id'] = txn['id']
                used_sheet.add(best)
                used_ynab.add(yi)

    # Pass 1: individual match — date variants + up to 3 days tolerance
    run_pass(3)

    # Pass 2: group match — multiple YNAB entries that sum to one spreadsheet row
    # (handles split bank transfers logged as a single invoice in the spreadsheet)
    unmatched_ynab = [(yi, ynab_txns[yi]) for yi in ynab_sorted if yi not in used_ynab]
    for si, row in enumerate(sheet_rows):
        if si in used_sheet: continue
        target = row['amount']
        for row_date in date_variants(row['date']):
            nearby = [(yi, t) for yi, t in unmatched_ynab
                      if abs(datetime.strptime(t['date'], '%Y-%m-%d').date() - row_date) <= timedelta(days=20)]
            found = False
            for size in range(2, min(6, len(nearby) + 1)):
                for combo in combinations(nearby, size):
                    total = sum(round(abs(t['amount']) / 1000, 2) for _, t in combo)
                    if abs(total - target) <= 0.02:
                        sheet_rows[si]['ynab_id'] = combo[0][1]['id']  # first ID as anchor
                        used_sheet.add(si)
                        for yi2, _ in combo:
                            used_ynab.add(yi2)
                        found = True
                        break
                if found: break
            if si in used_sheet: break

    # Add genuinely unmatched YNAB entries as new rows (skip cash withdrawals)
    for yi, txn in unmatched_ynab:
        if yi in used_ynab: continue
        payee = (txn.get('payee_name', '') or '').lower()
        if any(k in payee for k in SKIP_PAYEES): continue
        yd    = datetime.strptime(txn['date'], '%Y-%m-%d').date()
        pname = txn.get('payee_name', '') or ''
        memo  = txn.get('memo', '') or ''
        sheet_rows.append({
            'date':      yd,
            'amount':    round(abs(txn['amount']) / 1000, 2),
            'desc':      memo,
            'payee':     pname,
            'tag':       infer_tag(pname, memo),
            'fonte':     '',
            'ref':       '',
            'drive':     '',
            'ynab_id':   txn['id'],
            'ynab_only': True,
        })

    return sorted(sheet_rows, key=lambda r: r['date'])

# --- 4. Build new workbook ---

def build_workbook(rows):
    wb = openpyxl.Workbook()

    H_FONT    = Font(bold=True, color='FFFFFF')
    H_FILL    = PatternFill('solid', fgColor='1F497D')
    YNAB_FILL = PatternFill('solid', fgColor='FFF2CC')  # yellow = needs review

    # ── Sheet 1: Gastos ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = '💸 Gastos'

    COLS   = ['Data', 'Valor', 'Descrição', 'Pago a', 'Tag', 'Fonte', 'Método', 'YNAB_ID', 'Ref', 'Drive']
    WIDTHS = [12,     12,      35,           22,       28,    8,       8,        38,         22,    45]

    for j, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(1, j, h)
        c.font, c.fill = H_FONT, H_FILL
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'

    ynab_only_rows = []
    for row in rows:
        r = ws.max_row + 1
        ws.cell(r, 1, row['date'])
        ws.cell(r, 2, row['amount'])
        ws.cell(r, 3, row['desc'])
        ws.cell(r, 4, row['payee'])
        ws.cell(r, 5, row['tag'])
        ws.cell(r, 6, row['fonte'])
        ws.cell(r, 7, 'Banco' if row['ynab_only'] else '')
        ws.cell(r, 8, row['ynab_id'] or '')
        ws.cell(r, 9, row['ref'])
        ws.cell(r, 10, row['drive'])
        ws.cell(r, 1).number_format = 'DD/MM/YYYY'
        ws.cell(r, 2).number_format = '#,##0.00 €'
        if row['ynab_only']:
            for c in range(1, 11):
                ws.cell(r, c).fill = YNAB_FILL
            ynab_only_rows.append(row)

    last_row = ws.max_row

    # ── Sheet 2: Dashboard ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('📊 Dashboard')

    ws2['A1'] = 'CUSTOS CASA NOVA'
    ws2['A1'].font = Font(bold=True, size=14, color='1F497D')

    ws2['A3'] = 'Total gasto'
    ws2['B3'] = f"=SUM('💸 Gastos'!B2:B{last_row})"
    ws2['B3'].number_format = '#,##0.00 €'
    ws2['A3'].font = Font(bold=True)

    ws2['A4'] = 'Disponível (YNAB, hoje)'
    ws2['B4'] = YNAB_TOTAL
    ws2['B4'].number_format = '#,##0.00 €'
    ws2['A4'].font = Font(bold=True)

    ws2['A5'] = '  Banco — MG Ordenado'
    ws2['B5'] = YNAB_BANK
    ws2['B5'].number_format = '#,##0.00 €'

    ws2['A6'] = '  Cash — Cache obra'
    ws2['B6'] = YNAB_CASH
    ws2['B6'].number_format = '#,##0.00 €'

    ws2['A8'] = 'Gasto por categoria'
    ws2['A8'].font = Font(bold=True)
    ws2['B8'] = 'Total'
    ws2['B8'].font = Font(bold=True)
    ws2['C8'] = '%'
    ws2['C8'].font = Font(bold=True)

    tag_sums = defaultdict(float)
    for row in rows:
        if row['tag']:
            tag_sums[row['tag']] += row['amount']
    grand_total = sum(tag_sums.values())

    r = 9
    for tag, total in sorted(tag_sums.items(), key=lambda x: -x[1]):
        ws2.cell(r, 1, tag)
        ws2.cell(r, 2, round(total, 2))
        ws2.cell(r, 2).number_format = '#,##0.00 €'
        ws2.cell(r, 3, round(total / grand_total * 100, 1) if grand_total else 0)
        ws2.cell(r, 3).number_format = '0.0"%"'
        r += 1

    ws2['A3'].font = ws2['A4'].font = ws2['A8'].font = Font(bold=True)
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 8

    # ── Sheet 3: Para Rever ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet('🔍 Para Rever')
    ws3['A1'] = (f'{len(ynab_only_rows)} transações do banco (YNAB) sem correspondência '
                 f'no spreadsheet — preenche Fonte (LB/MG) e revê Tag')
    ws3['A1'].font = Font(bold=True)

    REC_COLS   = ['Data', 'Valor', 'YNAB Payee', 'Memo', 'Tag inferida', 'Fonte', 'YNAB ID']
    REC_WIDTHS = [12,     12,      45,             35,     28,             8,       38]
    for j, (h, w) in enumerate(zip(REC_COLS, REC_WIDTHS), 1):
        c = ws3.cell(2, j, h)
        c.font = Font(bold=True)
        ws3.column_dimensions[get_column_letter(j)].width = w

    for i, row in enumerate(ynab_only_rows, 3):
        ws3.cell(i, 1, row['date'])
        ws3.cell(i, 2, row['amount'])
        ws3.cell(i, 3, row['payee'])
        ws3.cell(i, 4, row['desc'])
        ws3.cell(i, 5, row['tag'])
        ws3.cell(i, 6, '')  # user fills Fonte
        ws3.cell(i, 7, row['ynab_id'])
        ws3.cell(i, 1).number_format = 'DD/MM/YYYY'
        ws3.cell(i, 2).number_format = '#,##0.00 €'

    wb.save(OUTPUT)
    return ynab_only_rows

# --- Main ---

if __name__ == '__main__':
    print("Fetching YNAB transactions...")
    ynab = fetch_ynab()
    print(f"  {len(ynab)} transactions in Gastos Casa Nova")

    print("Loading spreadsheet...")
    sheet = load_sheet()
    print(f"  {len(sheet)} rows parsed")

    print("Matching and merging...")
    merged = match_and_merge(sheet, ynab)

    matched    = sum(1 for r in merged if r['ynab_id'] and not r['ynab_only'])
    ynab_only  = sum(1 for r in merged if r['ynab_only'])
    sheet_only = sum(1 for r in merged if not r['ynab_id'])
    total      = sum(r['amount'] for r in merged)

    print(f"  Matched (in both):              {matched}")
    print(f"  YNAB only (added, yellow):      {ynab_only}")
    print(f"  Spreadsheet only (cash/manual): {sheet_only}")
    print(f"  Total rows in new file:         {len(merged)}")
    print(f"  Total spent (all entries):      €{total:,.2f}")

    print("Writing new spreadsheet...")
    build_workbook(merged)
    print(f"  ✓ Saved to: {OUTPUT}")
